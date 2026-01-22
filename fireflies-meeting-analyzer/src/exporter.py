"""
Export-Modul für Analyseergebnisse.
Unterstützt Excel, Markdown und JSON Export.
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
from dataclasses import asdict

from .config import settings
from .ai_analyzer import AnalysisResult, PainPoint, Question, Objection


class ReportExporter:
    """Exportiert Analyseergebnisse in verschiedene Formate."""

    def __init__(self, export_dir: str = None):
        """
        Initialisiert den Exporter.

        Args:
            export_dir: Verzeichnis für Exports (default: aus Settings)
        """
        self.export_dir = Path(export_dir or settings.export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def _generate_filename(self, prefix: str, extension: str) -> Path:
        """Generiert einen eindeutigen Dateinamen."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return self.export_dir / f"{prefix}_{timestamp}.{extension}"

    def export_to_json(self, result: AnalysisResult, filename: str = None) -> str:
        """
        Exportiert Ergebnisse als JSON.

        Args:
            result: Analyseergebnis
            filename: Optional: Benutzerdefinierter Dateiname

        Returns:
            Pfad zur exportierten Datei
        """
        filepath = Path(filename) if filename else self._generate_filename("analyse", "json")

        data = result.to_dict()

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return str(filepath)

    def export_to_markdown(self, result: AnalysisResult, filename: str = None) -> str:
        """
        Exportiert Ergebnisse als Markdown-Dokument.

        Args:
            result: Analyseergebnis
            filename: Optional: Benutzerdefinierter Dateiname

        Returns:
            Pfad zur exportierten Datei
        """
        filepath = Path(filename) if filename else self._generate_filename("analyse", "md")

        md_content = self._build_markdown(result)

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(md_content)

        return str(filepath)

    def _build_markdown(self, result: AnalysisResult) -> str:
        """Erstellt den Markdown-Inhalt."""
        lines = []

        # Header
        lines.append(f"# Meeting-Analyse Report")
        lines.append(f"")
        lines.append(f"**Erstellt:** {result.timestamp.strftime('%d.%m.%Y %H:%M')}")
        lines.append(f"**Analysierte Meetings:** {result.meetings_analyzed}")
        lines.append(f"**Lead-Aussagen:** {result.total_lead_statements}")
        lines.append("")
        lines.append("---")
        lines.append("")

        # Kunden-Profile
        if result.customer_profiles:
            lines.append("## 1. Kunden-Insight-Profile")
            lines.append("")
            for i, profile in enumerate(result.customer_profiles, 1):
                lines.append(f"### Profil {i}")
                if profile.company_name:
                    lines.append(f"- **Unternehmen:** {profile.company_name}")
                if profile.industry:
                    lines.append(f"- **Branche:** {profile.industry}")
                if profile.decision_maker_role:
                    lines.append(f"- **Entscheider-Rolle:** {profile.decision_maker_role}")
                if profile.business_context:
                    lines.append(f"- **Geschäftskontext:** {profile.business_context}")
                if profile.timeline_urgency:
                    lines.append(f"- **Dringlichkeit:** {profile.timeline_urgency}")
                if profile.supporting_quotes:
                    lines.append("")
                    lines.append("**Unterstützende Zitate:**")
                    for quote in profile.supporting_quotes:
                        lines.append(f"> \"{quote.get('quote', '')}\"")
                        if quote.get('context'):
                            lines.append(f"> *{quote.get('context')}*")
                lines.append("")

        # Pain Points
        if result.pain_points:
            lines.append("## 2. Pain Point Matrix")
            lines.append("")

            # Gruppiere nach Kategorie
            by_category: Dict[str, List[PainPoint]] = {}
            for pp in result.pain_points:
                cat = pp.category or "Sonstiges"
                if cat not in by_category:
                    by_category[cat] = []
                by_category[cat].append(pp)

            for category, points in by_category.items():
                lines.append(f"### {category}")
                lines.append("")
                for pp in points:
                    lines.append(f"**{pp.description}** (Priorität: {pp.impact_level})")
                    lines.append("")
                    lines.append(f"> \"{pp.direct_quote}\"")
                    lines.append(f"> — *{pp.speaker}*")
                    lines.append("")
                    if pp.context:
                        lines.append(f"**Kontext:** {pp.context}")
                    if pp.desired_outcome:
                        lines.append(f"**Gewünschtes Ergebnis:** {pp.desired_outcome}")
                    if pp.impact_statement:
                        lines.append(f"**Auswirkung:** {pp.impact_statement}")
                    lines.append("")

        # Fragen
        if result.questions:
            lines.append("## 3. Lead-Fragen")
            lines.append("")

            # Gruppiere nach Kategorie
            by_category: Dict[str, List[Question]] = {}
            for q in result.questions:
                cat = q.category or "Allgemein"
                if cat not in by_category:
                    by_category[cat] = []
                by_category[cat].append(q)

            for category, questions in by_category.items():
                lines.append(f"### {category}")
                lines.append("")
                for q in questions:
                    lines.append(f"**Frage:** \"{q.text}\"")
                    lines.append(f"- *{q.speaker}*")
                    if q.underlying_concern:
                        lines.append(f"- **Zugrundeliegende Sorge:** {q.underlying_concern}")
                    lines.append("")

        # Einwände
        if result.objections:
            lines.append("## 4. Einwand-Analyse")
            lines.append("")
            for obj in result.objections:
                lines.append(f"### \"{obj.objection_text}\"")
                lines.append("")
                lines.append(f"> \"{obj.direct_quote}\"")
                lines.append(f"> — *{obj.speaker}*")
                lines.append("")
                if obj.emotional_undertone:
                    lines.append(f"**Emotionale Färbung:** {obj.emotional_undertone}")
                if obj.root_cause:
                    lines.append(f"**Ursache:** {obj.root_cause}")
                if obj.resolution_pathway:
                    lines.append(f"**Lösungsweg:** {obj.resolution_pathway}")
                if obj.conversion_trigger:
                    lines.append(f"**Conversion Trigger:** {obj.conversion_trigger}")
                lines.append("")

        # Sprachanalyse
        if result.language_patterns:
            lines.append("## 5. Sprachanalyse")
            lines.append("")

            # Gruppiere nach Kategorie
            categories = {
                "industry_term": "Fachterminologie",
                "emotional": "Emotionale Sprache",
                "power_word": "Power Words",
                "metaphor": "Metaphern & Analogien"
            }

            for cat_key, cat_name in categories.items():
                patterns = [lp for lp in result.language_patterns if lp.category == cat_key]
                if patterns:
                    lines.append(f"### {cat_name}")
                    lines.append("")
                    for lp in patterns:
                        lines.append(f"- **\"{lp.phrase}\"** — {lp.speaker}")
                        if lp.context:
                            lines.append(f"  - Kontext: {lp.context}")
                    lines.append("")

        # Value Propositions
        if result.value_propositions:
            lines.append("## 6. Value Propositions")
            lines.append("")
            for vp in result.value_propositions:
                lines.append(f"**Kundenbedürfnis:** {vp.get('customer_need', '')}")
                lines.append(f"**Lösung:** {vp.get('solution_alignment', '')}")
                lines.append(f"**Nutzenaussage:** {vp.get('benefit_statement', '')}")
                lines.append("")

        # Marketing Empfehlungen
        if result.marketing_recommendations:
            lines.append("## 7. Marketing-Empfehlungen")
            lines.append("")
            for rec in result.marketing_recommendations:
                lines.append(f"### {rec.get('section', 'Allgemein')}")
                if rec.get('customer_quote'):
                    lines.append(f"> \"{rec.get('customer_quote')}\"")
                lines.append(f"**Empfehlung:** {rec.get('recommended_message', '')}")
                lines.append("")

        # Messaging Guidelines
        if result.messaging_guidelines:
            lines.append("## 8. Messaging-Leitfaden")
            lines.append("")
            for msg in result.messaging_guidelines:
                lines.append(f"**Begriff:** \"{msg.get('effective_term', '')}\"")
                if msg.get('customer_quote'):
                    lines.append(f"> \"{msg.get('customer_quote')}\"")
                lines.append(f"**Empfehlung:** {msg.get('usage_recommendation', '')}")
                lines.append("")

        # Conversion Triggers
        if result.conversion_triggers:
            lines.append("## 9. Conversion Triggers")
            lines.append("")
            for trigger in result.conversion_triggers:
                lines.append(f"- {trigger}")
            lines.append("")

        return "\n".join(lines)

    def export_to_excel(self, result: AnalysisResult, filename: str = None) -> str:
        """
        Exportiert Ergebnisse als Excel-Datei.

        Args:
            result: Analyseergebnis
            filename: Optional: Benutzerdefinierter Dateiname

        Returns:
            Pfad zur exportierten Datei
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise ImportError("pandas und openpyxl werden benötigt. Bitte installieren.")

        filepath = Path(filename) if filename else self._generate_filename("analyse", "xlsx")

        wb = Workbook()

        # Übersicht
        ws_overview = wb.active
        ws_overview.title = "Übersicht"
        ws_overview["A1"] = "Meeting-Analyse Report"
        ws_overview["A1"].font = Font(bold=True, size=14)
        ws_overview["A3"] = "Erstellt:"
        ws_overview["B3"] = result.timestamp.strftime("%d.%m.%Y %H:%M")
        ws_overview["A4"] = "Analysierte Meetings:"
        ws_overview["B4"] = result.meetings_analyzed
        ws_overview["A5"] = "Lead-Aussagen:"
        ws_overview["B5"] = result.total_lead_statements

        # Pain Points Sheet
        if result.pain_points:
            ws_pain = wb.create_sheet("Pain Points")
            headers = ["Kategorie", "Beschreibung", "Zitat", "Sprecher", "Priorität", "Kontext", "Gewünschtes Ergebnis"]
            ws_pain.append(headers)

            for cell in ws_pain[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            for pp in result.pain_points:
                ws_pain.append([
                    pp.category,
                    pp.description,
                    pp.direct_quote,
                    pp.speaker,
                    pp.impact_level,
                    pp.context,
                    pp.desired_outcome
                ])

        # Fragen Sheet
        if result.questions:
            ws_questions = wb.create_sheet("Fragen")
            headers = ["Frage", "Sprecher", "Kategorie", "Zugrundeliegende Sorge", "Kontext"]
            ws_questions.append(headers)

            for cell in ws_questions[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            for q in result.questions:
                ws_questions.append([
                    q.text,
                    q.speaker,
                    q.category,
                    q.underlying_concern,
                    q.context
                ])

        # Einwände Sheet
        if result.objections:
            ws_obj = wb.create_sheet("Einwände")
            headers = ["Einwand", "Zitat", "Sprecher", "Emotionale Färbung", "Ursache", "Lösungsweg", "Conversion Trigger"]
            ws_obj.append(headers)

            for cell in ws_obj[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            for obj in result.objections:
                ws_obj.append([
                    obj.objection_text,
                    obj.direct_quote,
                    obj.speaker,
                    obj.emotional_undertone,
                    obj.root_cause,
                    obj.resolution_pathway,
                    obj.conversion_trigger
                ])

        # Sprachanalyse Sheet
        if result.language_patterns:
            ws_lang = wb.create_sheet("Sprachanalyse")
            headers = ["Kategorie", "Ausdruck", "Sprecher", "Kontext"]
            ws_lang.append(headers)

            for cell in ws_lang[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            category_names = {
                "industry_term": "Fachterminologie",
                "emotional": "Emotionale Sprache",
                "power_word": "Power Words",
                "metaphor": "Metaphern"
            }

            for lp in result.language_patterns:
                ws_lang.append([
                    category_names.get(lp.category, lp.category),
                    lp.phrase,
                    lp.speaker,
                    lp.context
                ])

        wb.save(filepath)
        return str(filepath)
